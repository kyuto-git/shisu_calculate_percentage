# ythonのライブラリをインポートし、
import pandas as pd
import os
# このライブラリはExcelファイルの読み書きに使われます。
# 「openpyxl.styles」からNamedStyleクラスをインポートしています。これはExcelのセルのスタイルを定義するために使われます。
import openpyxl as px
from openpyxl.styles import NamedStyle, Font, PatternFill


def calculate_rates(df):
    # Calculate 4着内率 and 5着内率 and 1・3着率 and 2・3着率
    df["4着内率"] = ((df["１着数"] + df["２着数"] + df["３着数"] + df["４着数"]) / df["総データ数"])
    df["5着内率"] = ((df["１着数"] + df["２着数"] + df["３着数"] + df["４着数"] + df["５着数"]) / df["総データ数"])
    df["1・3着率"] = ((df["１着数"] + df["３着数"]) / df["総データ数"])
    df["2・3着率"] = ((df["２着数"] + df["３着数"]) / df["総データ数"])

    # 「[単勝]回収値実現平均オッズ」を計算。「勝率」列の値からパーセンテージ記号を削除し、数値に変換して100で割っています。これにより実数値の勝率が計算されます。
    tansho = df["勝率"].str.replace('%', '').astype(float) / 100
    fukusho = df["複勝率"].str.replace('%', '').astype(float) / 100
    df['[単勝]回収値実現平均オッズ'] = (((df['単勝適正回収値'] * 10) / 1000) / tansho).round(1)
    df['[複勝]回収値実現平均オッズ'] = (((df['複勝回収値'] * 10) / 1000) / fukusho).round(1)
    return df

# 条件に基づいて色を変更する関数
def apply_color_conditions(writer):
     # 各シートに対してスタイルを適用
    for sheet in writer.sheetnames:
        sheet = writer[sheet]

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column):
            # E, H, K列の値が0.35以上の場合、文字色を赤にする
            for col_idx in [4, 7, 10]:  # E, H, K列のインデックス
                if row[col_idx].value is not None and row[col_idx].value >= 0.35:
                    row[col_idx].font = Font(color="FF0000")

            # M列の複数の条件に基づいて、セルの色を変更する
            m_val = row[12].value  # M列
            g_val = row[6].value   # G列
            n_val = row[13].value  # N列
            v_val = row[21].value  # V列
            if m_val is not None:
                # 条件に基づいて色を設定
                color = None
                if m_val >= 0.9:
                    color = "00FF33"  # 緑色
                elif (m_val >= 0.75 and (g_val >= 80 or n_val >= 80)) or (m_val >= 0.6 and g_val >= 100 and n_val >= 100):
                    color = "FFFF00"  # 黄色
                elif (m_val >= 0.5 and (g_val >= 80 or n_val >= 80) and v_val >= 0.6) or \
                    (m_val >= 0.5 and (g_val >= 120 or n_val >= 120) and v_val >= 0.5):
                    color = "FF9900"  # オレンジ色
                elif (m_val >= 0.35 and (g_val >= 80 or n_val >= 80) and v_val >= 0.5) or \
                    (m_val >= 0.35 and (g_val >= 120 or n_val >= 120) and v_val >= 0.4):
                    color = "00FFFF"  # シアン色
                elif m_val >= 0.25 and g_val >= 150 and n_val >= 150 and v_val >= 0.35:
                    color = "999999"  # グレー色
                elif m_val == 0 or (m_val < 0.1 and (g_val < 100 or n_val < 100)):
                    color = "000000"  # 黒色

                if color:
                    row[12].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            pass  # この部分を具体的な条件に基づいて色を変更するコードで置き換える


# この関数は3つの引数を受け取ります：writer（Excelファイルを書き出すためのオブジェクト）、original_df（元のDataFrame）、そしてpercentage_cols（パーセンテージ形式にするべき列のリスト）。
def percentage_format(writer, original_df, percentage_cols):
    # NamedStyleオブジェクトを作成し、Excelのセルのスタイルを定義しています。
    percentage_style = NamedStyle(
        # スタイルの名前を「Percentage」、NamedStyleオブジェクトにパーセンテージ形式（小数点以下2桁）を設定しています。
        name='Percentage',
        number_format='0.00%'
    )
    # writer.sheetsの各シートに対してループを実行しています。ここでのsheet_nameはシート名を表し、dfはそのシートのDataFrameです。
    for sheet_name, df in writer.sheets.items():
        # シート名がoriginal_dfに含まれていない場合、ループを続行するためにcontinueを使用しています。
        if sheet_name not in original_df:
            continue
        # シート名に基づいて正しいDataFrameをoriginal_dfから取得しています。
        df = original_df[sheet_name]
        # シート名に基づいてExcelのシートを取得しています。
        sheet = writer.sheets[sheet_name]
        # DataFrameの各列に対してループを実行しています。idxは列のインデックス、colは列名です。
        for idx, col in enumerate(df.columns):
            # Excelは1から始まるインデックスを使用するため、列インデックスを1に調整しています。
            col_idx = idx + 1
            # 現在の列がパーセンテージ形式にするべき列のリストに含まれている場合、特定の処理を行うためにこの条件分岐を使用しています。
            if col in percentage_cols:
                # このループはDataFrameの各行に対して実行されます。range(2, len(df) + 2)は、1から始まるインデックスを使用し、ヘッダー行をスキップするために2から開始しています。
                for row_idx in range(2, len(df) + 2):
                    # Excelのシートから特定のセルを取得しています。ここでのrow_idxとcol_idxはそれぞれ行と列のインデックスです。
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    # 取得したセルに先ほど定義したパーセンテージ形式のスタイルを適用しています。
                    cell.style = percentage_style


def process_csv_files():
    csv_dir = 'csv_files'
    output_dir = 'output'
    output_file = 'compiled_data.xlsx'

    # 出力ディレクトリが存在しない場合は作成します。exist_ok=Trueは、ディレクトリが既に存在する場合にエラーを発生させないためのオプションです。
    os.makedirs(output_dir, exist_ok=True)

    # 指定されたディレクトリ内の.csvで終わる全てのファイルをリストとして取得しています。
    csv_files = [f for f in os.listdir(csv_dir) if f.endswith('.csv')]

    # PandasのExcelライターを作成し、writerに代入しています。ここではopenpyxlエンジンを使用しています。
    writer = pd.ExcelWriter(os.path.join(output_dir, output_file), engine='openpyxl')

    percentage_cols = [
        "勝率", "２着率", "連対率", "３着率", "複勝率", "１人気率",
        "３人気内率", "５人気内率", "６人気以下率", "4着内率", "5着内率", "1・3着率", "2・3着率"
    ]

    # 空の辞書df_dictを作成しています。
    df_dict = {}
    
    # csv_filesリスト内の各CSVファイルに対してループを実行しています。
    for csv_file in csv_files:
        # CSVファイルを読み込んでDataFrame(df)に格納しています。ここでのencoding='shift_jis'は、日本語のエンコーディングに対応するために使用されています。
        df = pd.read_csv(os.path.join(csv_dir, csv_file), encoding='shift_jis')

        # calculate_rates関数を使用してDataFrameに対する計算を実施しています。これにより、特定の統計的な値が計算されます。
        df = calculate_rates(df)

        # 各CSVファイルを処理するループ内で、`df_dict` 辞書にデータフレームを追加しました。これにより、シート名をキーとして、対応するデータフレームを辞書に保存します。
        sheet_name = os.path.splitext(csv_file)[0]
        df_dict[sheet_name] = df
        
        # DataFrameをExcelファイルに書き出しています。index=Falseは、インデックスを出力しないためのオプションです。
        df.to_excel(writer, sheet_name=os.path.splitext(csv_file)[0], index=False)

    # # 指定されたセルや文字の色の変更を行う
    # apply_color_conditions(writer)

    # # 先ほど定義したpercentage_format関数を使用して、Excelシートにパーセンテージ形式を適用しています。この関数は先に定義されたwriter、df_dict（シート名に基づくDataFrameの辞書）、およびpercentage_cols（パーセンテージ形式に変換する列のリスト）を引数に取ります。
    percentage_format(writer, df_dict, percentage_cols)

    # writerオブジェクトを閉じて、Excelファイルの書き出しを完了します。
    writer.close()

if __name__ == "__main__":
    process_csv_files()
